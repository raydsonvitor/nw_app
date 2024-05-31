from customtkinter import *
from tkinter.ttk import Spinbox
from CTkSpinbox import *
import openpyxl as op
from datetime import date
from CTkMessagebox import CTkMessagebox
from time import sleep
from defs_2_0 import *


def Futrica():
    wb=op.load_workbook(r'excell\nw_barbearia_2024.xlsx')
    ws=wb['04-24']

    for item in range(ws.max_row-1):
        celula = ws.cell(row=item+2, column=3)
        valor = celula.value
        if valor == 'VTR':
            valor = 'VITOR'
        elif valor == 'LF':
            valor = 'FERNANDO'
        celula.value = valor
        print(celula.value)    

    wb.save(r'excell\nw_barbearia_2024.xlsx')

def get_bebidas_fat_semanal(periodo):
    wb = op.load_workbook(r'excell\nw_barbearia_2024.xlsx')
    ws = wb[periodo]
    rows = list(ws.values)[1:]
    lista_datas = Get_lista_datas_semana()
    total = 0
    for row in rows:
        if row[1] in lista_datas:
            if row[4] != None:#filtrando as rows que tem registro de bebida
                if '+' in row[7]:#tratando os componentes
                    row=list(row)
                    row[7] =Soma(row[7].split(' + '))
                total+=float(row[7])
                print(row[1], row[7])
    
    return total

def Check_datas(data):
    lista_datas = Get_lista_datas_semana()
    if data in lista_datas:
        return True
    else:
        return False


def Update_lista_datas_semana(data, ano):
    def Get_first_week_date(data, ano, weekday_indice):#identificar da data da ultima segunda-feira seja ela no mes ou ano passado
        try:
            data_splited = data.split('-')
            day = int(data_splited[0])
            month = int(data_splited[1])
            year = int(ano)
            if weekday_indice == 0:
                return data
            else:#para o caso de o dia em questão não for segunda
                day = int(day)
                month = int(month)
                for i in range(weekday_indice+1):
                    if day < 1:
                        if month == 1:
                            day = 31
                            month = 12
                            ano = str(int(ano)-1)
                        else:    
                            day = DiasMes(month-1, ano)
                            month -= 1
                    if i == weekday_indice:
                        return [int(ano), month, day]
                    day -= 1
        except:
            print('Erro na subfunção Get_first_week_date da função Update_lista_semanal, arq defs.py.')
    try:
        day, month = data.split('-')
        data_datetime = date(int(ano),int(month),int(day))
        weekday_indice = data_datetime.weekday()
        if weekday_indice != 0:
            year, month, day = Get_first_week_date(data, ano, weekday_indice)
        else:
            day = data_datetime.day
            month = data_datetime.month
            year = data_datetime.year
        #formar a lista de datas desde a ultima segunda feira
        lista_datas = []
        dias_mes = DiasMes(month, ano[2:])
        for i in range(7):#loop que acrescenta as datas da semana na lista
            if day > dias_mes:#condição que recunha a virada do mes
                day = 1
                if month < 12:
                    month += 1
                else:
                    month=1
                    year += 1
            variable_data = Zero_adder(str(day))+'-'+Zero_adder(str(month))
            lista_datas.append(variable_data)
            day += 1
        with open('txts\datas_semana.txt', 'w') as a:
            joined_list=';'.join(lista_datas)
            a.write(joined_list)
            print(f'Lista de datas da semana foram atualizadas: {lista_datas}')
    except:
        print('Erro na função update_datas_semana no arq def.py.')

def Obter_faturamento_semanal_by_barbeiro(data, periodo):
    def Get_meses_lista_datas_semana(lista_datas):
        #formar lista dos meses diferentes separando-os dos dias e tratando-os
        meses = []
        last_mes = ''
        for data in lista_datas:
            data_splited = data.split('-')
            mes = data_splited[1]
            if last_mes != mes:
                meses.append(mes)
            last_mes = mes
        return meses # [atual] | [antigo, atual]

    #definições importantes
    data_splited = data.split('-')
    mes = data_splited[1]
    periodo_splited = periodo.split('-')
    ano = '20' + periodo_splited[1]
    ano_abrev = periodo_splited[1]
    #obtenção das datas da semana
    if Check_datas(data) == False:#check se as datas da semana são válidas para a data atual
        Update_lista_datas_semana(data, ano)#atualiza as datas da semana
        lista_datas = Get_lista_datas_semana()
    else:
        lista_datas = Get_lista_datas_semana() 
    #calculando o faturamento de cada profissional no(s) periodo(s)
    print(f'peguei a lista de datas:{lista_datas}')
    lista_meses = Get_meses_lista_datas_semana(lista_datas)#o len dessa lista def se as datas da semana estão contidos somente no periodo mensal atual ou se tem mais de um
    print(f'lista meses das datas da semana: {lista_meses}')
    profissional1 = profissional2 = profissional3 = profissional4 = 0
    for i, mes in enumerate(lista_meses[::-1], start=1):#passando pelo(s) database(s) do mais novo pro mais velho
        print(f'analisando o mes: {mes}')
        if i == 2:#analisando apenas o segundo mes do for
            if mes == '12':
                ano = str(int(ano)-1)
                print('troquei o ano abrev')
        #abrir o data base do periodo mensal da instancia
        wb = op.load_workbook(fr'excell\nw_barbearia_{ano}.xlsx')
        ws = wb[f'{mes}-{ano_abrev}']
        rows = list(ws.values)
        for row in rows:
            #filtrando data
            if row[1] in lista_datas:
                row = list(row)
                if '+' in row[7]:
                    row[7]=Soma(row[7].split(' + '))
                #filtrando o profissional
                if row[2] == profissionais[0]:
                    profissional1 += float(row[7])
                elif row[2] == profissionais[1]:
                    profissional2 += float(row[7])
                elif row[2] == profissionais[2]:
                    profissional3 += float(row[7])
                elif row[2] == profissionais[3]:
                    profissional4 += float(row[7])
    return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']




    '''   #getting lista das rows
    sheet = wb[periodo]
    rows = list(sheet.values)
    #identificando as rows certas
    profissional1 = profissional2 = profissional3 = profissional4 = 0
    for row in rows:
        #tratando a lista
        row = list(row)
        if '+' in row[7]:
            row[7]=Soma(row[7].split(' + '))
        #filtrando data
        if row[1] in lista_datas:
            #filtrando o profissional
            if row[2] == profissionais[0]:
                profissional1 += float(row[7])
            elif row[2] == profissionais[1]:
                profissional2 += float(row[7])
            elif row[2] == profissionais[2]:
                profissional3 += float(row[7])
            elif row[2] == profissionais[3]:
                profissional4 += float(row[7])
    return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']
else:#True quer dizer que todas as datas NÃO estão no mesmo periodo
    if mes == '01':#Para o caso de o MES anterior estar contido no ANO anterior
        mes_anterior = '12'
        ano_anterior_abrev = str(int(ano[2:])-1)
        #ws mes anterior
        ws = wb[f'{Zero_adder(mes_anterior)}-{ano_anterior_abrev}']
    else:
        mes_anterior = str(int(mes)-1)
        ano_anterior_abrev = str(int(ano[2:])-1)
        #ws mes anterior
        ws = wb[f'{Zero_adder(mes_anterior)}-{ano_abrev}']
    #pegando a lista de datas da semana
    lista_datas = Get_lista_datas_semana()
    rows = list(ws.values)
    #identificando as rows certas
    profissional1 = profissional2 = profissional3 = profissional4 = 0
    for row in rows:#periodo anterior
        #filtrando data
        if row[1] in lista_datas:
            row = list(row)
            if '+' in row[7]:
                row[7]=Soma(row[7].split(' + '))
            #filtrando o profissional
            if row[2] == profissionais[0]:
                profissional1 += float(row[7])
            elif row[2] == profissionais[1]:
                profissional2 += float(row[7])
            elif row[2] == profissionais[2]:
                profissional3 += float(row[7])
            elif row[2] == profissionais[3]:
                profissional4 += float(row[7])
    ano_abrev = ano[2:]
    #ws mes atual
    ws = wb[f'{Zero_adder(mes)}-{ano_abrev}']
    rows = list(sheet.values)
    for row in rows:#periodo atual
        #filtrando data
        if row[1] in lista_datas:
            row = list(row)
            if '+' in row[7]:
                row[7]=Soma(row[7].split(' + '))
            #filtrando o profissional
            if row[2] == profissionais[0]:
                profissional1 += float(row[7])
            elif row[2] == profissionais[1]:
                profissional2 += float(row[7])
            elif row[2] == profissionais[2]:
                profissional3 += float(row[7])
            elif row[2] == profissionais[3]:
                profissional4 += float(row[7])
    print([f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}'])
    return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']'''

    #print(f'Erro durante a execução da função Obter_faturamento_por_barbeiro_semanal no arquivo {arquivo_name}. lista retornada: [''0.00'', ''0.00'', ''0.00'', ''0.00'']')
    #return ['0.00', '0.00', '0.00', '0.00']


#x = Get_meses_lista_datas_semana(Get_lista_datas_semana())
