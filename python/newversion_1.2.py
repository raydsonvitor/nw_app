from tkinter import *
from tkinter.ttk import *
from defs import *
from customtkinter import *
import openpyxl as op
from PIL import Image

#MELHORIAS VERSAO 1.2: 
#caminhos\path dos arquivos ajustados para ambiente virtual e criação de um app.exe
#adição da tabview
#adição do programa abrir do tamanho padrao maximizado (geometria responsiva)
#exclusão do botao bebidas

janela_main = CTk()#janela_main=frame da tab 2
janela_main.state('zoomed')
janela_main.title('Night Wolf ADM')
janela_main.resizable(False, False)

#definicoes importantes
#cor das figuras dos botoes: #004aad

set_appearance_mode('dark')
set_default_color_theme('dark-blue')
fonte = CTkFont('arial', 15)
fonte_b = CTkFont('arial', 15, 'bold')
fonte_c = CTkFont('arial', 25, 'bold')
profissionais = ObterListaProfissionais()
despesas = ['CONSUMÍVEIS', 'ALUGUEL', 'ÁGUA', 'LUZ', 'INTERNET','FUNDOS', 'ESTORNO', 'OUTRO']
form_pgmt_entrada = ['DINHEIRO', 'PIX', 'DÉBITO', 'CRÉDITO','MENSAL']
form_pgmt_saida = ['DINHEIRO', 'BANCO', 'CARTÃO']
form_pgmt_cartao = ['DÉBITO', 'CRÉDITO']
colunas_database = ('id' ,'data', 'profissional' ,'histórico' ,'form pgmt' ,'entrada' ,'saída', 'hora')
periodo = GetPeriodo(colunas_database)
data = GetData()
ano = '20'+periodo[3:]

#maior: FRAME 8

#print info de inicialialzação
Lin()
print('Dados de inicialização\n')
print(f'data definida: {data}')
print(f'Periodo definido: {periodo}')
print('versao: 1.1')
Lin()

#Funções

def RegistrarCorte(periodo):
    try:
        #get row into excel
        profissional = combobox_barbeiro.get()
        serviço = []
        if check1.get() == 'CORTE':
            serviço.append(check1.get())
        if check2.get() == 'BARBA':
            serviço.append(check2.get())
        if check3.get() == 'SOBRANCELHA':
            serviço.append(check3.get())
        if check4.get() == 'PIGMENTAÇÃO':
            serviço.append(check4.get())
        if check5.get() == 'LUZES':
            serviço.append(check5.get())
        if check6.get() == 'PLATINADO':
            serviço.append(check6.get())
        servico = '+'.join(serviço)
        form_pgmt =  forma_pgmt.get()
        valor = entry_valor.get().strip().replace(',', '.')
        hora = GetHora()
        check = Check_0(profissional, servico, form_pgmt, valor)
        print(f'profissional > {profissional}, serviço >  {servico}, form. pgtm. > {form_pgmt}, valor > {valor}, hora > {hora}')
        if check == True:
            lastid = GetLastId(rf'excell\nw_barbearia_{ano}.xlsx', periodo, colunas_database)
            id = int(lastid)+1
            #getting row into dt
            book = op.load_workbook(rf'excell\nw_barbearia_{ano}.xlsx')
            sheet = book[periodo]
            sheet.append([id, data, profissional, servico ,form_pgmt, valor, '', hora])
            book.save(rf'excell\nw_barbearia_{ano}.xlsx')
            print(f'row salva no excell arq: nw_barbearia_{ano}.xlsx, sheet: {periodo}')
            #getting row into treeview
            tv.insert('', 'end', values=(id, data, profissional, servico ,form_pgmt, 'R$'+valor, '-', hora))
            #clearing widgets
            entry_valor.delete(0, 'end')
            combobox_barbeiro.set('PROFISSO.')
            forma_pgmt.set('FORM. PGMT.')
            check1_control.set('')
            check2_control.set('')
            check3_control.set('')
            check4_control.set('')
            check5_control.set('')
            check6_control.set('')
        else:
            print('Dados da area de registro incorretos')
        Atualizar()
    except:
        print('Ocorreu um erro  na função Registrar Corte')

def RegistrarSaida():
    ano = '20'+periodo[3:]
    #getting dados para a row
    lastid = GetLastId(rf'excell\nw_barbearia_{ano}.xlsx', periodo, colunas_database)
    id = int(lastid)+1
    despesa = frame_4_widget_1.get()
    form_pgmt = frame_4_widget_3.get()
    valor = frame_4_widget_5.get().strip().replace(',', '.')
    hora = GetHora()
    if Check_1(despesa, form_pgmt, valor)==True:
        #inserindo na treeviews
        row = [id, data, '', despesa, form_pgmt, '', valor, hora]
        tv.insert('', 'end', values=row)
        #inserindo no excell nw_barbearia_23
        book = op.load_workbook(rf'excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        sheet.append(row)
        book.save(rf'excell\nw_barbearia_{ano}.xlsx')
        print(f'row salva no excell arq: nw_barbearia_{ano}.xlsx sheet: {periodo}')
        #resetando widgets
        frame_4_widget_1.set('DESPESA')
        frame_4_widget_3.set('FORM. PGMT.')
        frame_4_widget_5.delete(0, 'end')
    else:
        print('Dados da area de registro incorretos')

    Atualizar()

def LoadData():
    try:
        ano = '20'+periodo[3:]
        book = op.load_workbook(rf'excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        list_values = list(sheet.values)
        if list_values != [colunas_database]:
            for col_name in list_values[0]:
                tv.heading(col_name, text= col_name)
            for row in list_values[1:]:
                if list(row)[1]==data:
                    row = list(row)
                    if row[-2] == None:
                        row[-2]='-'
                    else:
                        row[-2]=f'R${row[-2]}'
                    if row[3]==None:
                        row[3]='-'
                    if row[-2]==None:
                        row[-2]='-'
                    else:
                        row[-3]=f'R${row[-3]}'
                    tv.insert('', END,values=row)
            print(f'Load do arquvio nw_barbearia_{ano}.xlsx(sheet: {periodo}) para a Treewview realizado')
        else:
            print('Sheet vazia. Nada foi loadado na Treeview.')
    except:
        print('Erro ao dar o Load dos dados na treeview')

def Botao_bebida_window():
    janela_bebida = CTk(fg_color='white')
    janela_bebida.title('Registro Bebidas')
    janela_bebida.geometry('300x300+500+200')

    janela_bebida.mainloop()

def Atualizar():

    faturamento_por_barbeiro_diario = Obter_faturamento_por_barbeiro_diario(data, periodo)
    total_diario = Soma(faturamento_por_barbeiro_diario)
    label_1_textvar.set(f'R${faturamento_por_barbeiro_diario[0]}')
    label_3_textvar.set(f'R${faturamento_por_barbeiro_diario[1]}')
    label_5_textvar.set(f'R${faturamento_por_barbeiro_diario[2]}')
    label_7_textvar.set(f'R${faturamento_por_barbeiro_diario[3]}')
    label_9_textvar.set(f'R${total_diario:.2f}')

    faturamento_por_barbeiro_semanal = Obter_faturamento_por_barbeiro_semanal(data, periodo)
    total_semanal = Soma(faturamento_por_barbeiro_semanal)
    Frame_5_Widget_2_textvar.set(f'R${faturamento_por_barbeiro_semanal[0]}')
    Frame_5_Widget_4_textvar.set(f'R${faturamento_por_barbeiro_semanal[1]}')
    Frame_5_Widget_6_textvar.set(f'R${faturamento_por_barbeiro_semanal[2]}')
    Frame_5_Widget_8_textvar.set(f'R${faturamento_por_barbeiro_semanal[3]}')
    Frame_5_Widget_10_textvar.set(f'R${total_semanal:.2f}')

    faturamento_por_barbeiro_mensal = Obter_faturamento_por_barbeiro_mensal(periodo)
    total_mensal = Soma(faturamento_por_barbeiro_mensal)
    Frame_6_Widget_2_textvar.set(f'R${faturamento_por_barbeiro_mensal[0]}')
    Frame_6_Widget_4_textvar.set(f'R${faturamento_por_barbeiro_mensal[1]}')
    Frame_6_Widget_6_textvar.set(f'R${faturamento_por_barbeiro_mensal[2]}')
    Frame_6_Widget_8_textvar.set(f'R${faturamento_por_barbeiro_mensal[3]}')
    Frame_6_Widget_10_textvar.set(f'R${total_mensal:.2f}')

    faturamento_dia_by_formpgmt = Get_faturamento_dia_by_formpgmt(ano, periodo, data) # [dinheiro, cartão, pix]
    caixa = Obter_caixa(data, periodo)
    frame_7_Widget_1_textvar.set(f'R${faturamento_dia_by_formpgmt[0]}')
    frame_7_Widget_3_textvar.set(f'R${faturamento_dia_by_formpgmt[1]}')
    frame_7_Widget_5_textvar.set(f'R${faturamento_dia_by_formpgmt[2]}')
    frame_7_Widget_12_textvar.set(f'R${faturamento_dia_by_formpgmt[3]}')
    frame_7_Widget_8_textvar.set(f'R${caixa}')

    print('widgets atualizados')

    DeleteTreeviewData()
    LoadData()

    print('treeview linhas atualizadas')

def FecharCaixa(caixa_restante):
    try:
        #reescrevendo o arquivo
        with open(r'C:\Users\Customer\Desktop\nw_codigos\app\txts\caixa.txt', 'w') as a:
            a.write(caixa_restante)
            print(f'Saldo do caixa atualizado de {caixa} para {caixa_restante}')
        #limpando widget
        frame_7_Widget_10.delete(0, 'end')
    except:
        print('Erro na função FecharCaixa no arquivo app_barbearia.py')

def DeleteTreeviewData():
    try:
        for i in tv.get_children():
            tv.delete(i)
    except:
        print('Erro na função DeleteTreeviewDados, arquivo app.py')

def DeleteTreeviewItem():
    try:
        #pegando a row da treeview
        item= tv.selection()[0]
        valores=tv.item(item, 'values')
        tv.delete(item)
        id=valores[0]
        print(f'linha de código {id} deletado na treeview')
        #abrindo o database
        wb=op.load_workbook(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')
        ws=wb[periodo]
        rows=list(ws.values)[1:]
        index = 2
        for row in rows:
            if str(row[0]) == str(id):
                #check = Get_password_operator()
                #if check == True:
                ws.delete_rows(index, 1)
                wb.save(rf'C:\Users\Customer\Desktop\nw_codigos\app\excell\nw_barbearia_{ano}.xlsx')    
                print(f'linha de código {id} e index {index} deletado na database')
                break
            index+=1
        Atualizar()
    except:
        print('Erro na função DeleteTreeviewItem, arquivo app.py')
#tabview
tabview = CTkTabview(janela_main, width= 1366, height=700, border_width=1, border_color='white')
tabview.pack()
tab_principal = tabview.add('principal')
tab_pagamentos = tabview.add('pagamentos')

#  FRAME 0 \\ Upper Label

frame_0 = CTkFrame(tab_principal, width=1366, height=50, fg_color='blue')
frame_0.place(x=0, y=0)

label_0_1 = CTkLabel(frame_0, text=f'Data: {data}', font=CTkFont(size=30, weight='bold'))
label_0_1.place(x=570, y=6)

imagem_lobo = CTkImage(light_image=Image.open(r'C:\Users\Customer\Desktop\nw_codigos\app\images\imagem_lobo.png'), size=(45,45))
imagem_atualizar = CTkImage(light_image=Image.open(r'C:\Users\Customer\Desktop\nw_codigos\app\images\atualizar.png'), size=(30, 30 ))
imagem_excluir = CTkImage(light_image=Image.open(r'C:\Users\Customer\Desktop\nw_codigos\app\images\excluir.png'), size=(30, 30 ))

frame_0_widget_1 = CTkLabel(frame_0, image=imagem_lobo, text='')
frame_0_widget_1.place(x=10,y=3)
frame_0_widget_2 = CTkButton(frame_0, width=10, height=10 ,image=imagem_atualizar, text='', hover=True, fg_color='white', command=lambda:Atualizar())
frame_0_widget_2.place(x=1316, y=5)

#  FRAME 1  \\ Expositor faturamentos Diário

if True:
    #Expositor Faturamento Diário

    #definições pré-widget
    faturamento_por_barbeiro_diario = Obter_faturamento_por_barbeiro_diario(data, periodo)
    total_diario = Soma(faturamento_por_barbeiro_diario)
    label_1_textvar = StringVar(value=f'R${faturamento_por_barbeiro_diario[0]}')
    label_3_textvar = StringVar(value=f'R${faturamento_por_barbeiro_diario[1]}')
    label_5_textvar = StringVar(value=f'R${faturamento_por_barbeiro_diario[2]}')
    label_7_textvar = StringVar(value=f'R${faturamento_por_barbeiro_diario[3]}')
    label_9_textvar = StringVar(value=f'R${total_diario:.2f}')

    frame_1 = CTkFrame(tab_principal, width=201, height=180,border_width=1, border_color='white')
    frame_1.place(x=10, y=55)

    label_00 = CTkLabel(frame_1, width=197 ,text='Faturamento Diário', font=fonte_b, fg_color='blue')
    label_00.place(x=2, y=2)

    label_0 = CTkLabel(frame_1, width=98 ,text=profissionais[0], font=fonte_b, fg_color='black')
    label_0.place(x=1, y=30)
    label_1 = CTkLabel(frame_1, textvariable=label_1_textvar ,width=98 ,font=fonte_b, fg_color='black')
    label_1.place(x=101, y=30)

    label_2 = CTkLabel(frame_1, width=98 ,text=profissionais[1], font= fonte_b, fg_color='black')
    label_2.place(x=1, y=60)
    label_3 = CTkLabel(frame_1, textvariable=label_3_textvar ,width=98, font= fonte_b, fg_color='black') 
    label_3.place(x=101, y=60)

    label_4 = CTkLabel(frame_1, width=98 ,text=profissionais[2], font= fonte_b, fg_color='black')
    label_4.place(x=1, y=90)
    label_5 = CTkLabel(frame_1, textvariable=label_5_textvar ,width=98 , font= fonte_b, fg_color='black') 
    label_5.place(x=101, y=90)
    
    label_6 = CTkLabel(frame_1, width=98 ,text=profissionais[3], font= fonte_b, fg_color='black')
    label_6.place(x=1, y=120)
    label_7 = CTkLabel(frame_1, textvariable=label_7_textvar , width=98 , font= fonte_b, fg_color='black') 
    label_7.place(x=101, y=120)

    label_8 = CTkLabel(frame_1, width=98 ,text='TOTAL', font= fonte_b, fg_color='white', text_color='black')
    label_8.place(x=1, y=150)
    label_9 = CTkLabel(frame_1, textvariable=label_9_textvar , width=98 , font= fonte_b, fg_color='white', text_color='black') 
    label_9.place(x=101, y=150)

#   FRAME 5 \\ Expositor faturamentos Semanal

if True:
    #Expositor Faturamento semanal

    #definições pré-widget
    faturamento_por_barbeiro_semanal = Obter_faturamento_por_barbeiro_semanal(data, periodo)
    total_semanal = Soma(faturamento_por_barbeiro_semanal)
    Frame_5_Widget_2_textvar = StringVar(value=f'R${faturamento_por_barbeiro_semanal[0]}')
    Frame_5_Widget_4_textvar = StringVar(value=f'R${faturamento_por_barbeiro_semanal[1]}')
    Frame_5_Widget_6_textvar = StringVar(value=f'R${faturamento_por_barbeiro_semanal[2]}')
    Frame_5_Widget_8_textvar = StringVar(value=f'R${faturamento_por_barbeiro_semanal[3]}')
    Frame_5_Widget_10_textvar = StringVar(value=f'R${total_semanal:.2f}')

    frame_5 = CTkFrame(tab_principal, width=201, height=180,border_width=1, border_color='white')
    frame_5.place(x=10, y=255)

    Frame_5_Widget_0 = CTkLabel(frame_5, width=197 ,text='Faturamento Semanal', font=fonte_b, fg_color='blue')
    Frame_5_Widget_0.place(x=2, y=2)

    Frame_5_Widget_1 = CTkLabel(frame_5, width=98 ,text=profissionais[0], font=fonte_b, fg_color='black')
    Frame_5_Widget_1.place(x=1, y=30)
    Frame_5_Widget_2 = CTkLabel(frame_5, textvariable=Frame_5_Widget_2_textvar ,width=98 ,font=fonte_b, fg_color='black')
    Frame_5_Widget_2.place(x=101, y=30)

    Frame_5_Widget_3 = CTkLabel(frame_5, width=98 ,text=profissionais[1], font= fonte_b, fg_color='black')
    Frame_5_Widget_3    .place(x=1, y=60)
    Frame_5_Widget_4 = CTkLabel(frame_5, textvariable=Frame_5_Widget_4_textvar ,width=98, font= fonte_b, fg_color='black') 
    Frame_5_Widget_4.place(x=101, y=60)

    Frame_5_Widget_5 = CTkLabel(frame_5, width=98 ,text=profissionais[2], font= fonte_b, fg_color='black')
    Frame_5_Widget_5.place(x=1, y=90)
    Frame_5_Widget_6 = CTkLabel(frame_5, textvariable=Frame_5_Widget_6_textvar ,width=98 , font= fonte_b, fg_color='black') 
    Frame_5_Widget_6.place(x=101, y=90)
    
    Frame_5_Widget_7 = CTkLabel(frame_5, width=98 ,text=profissionais[3], font= fonte_b, fg_color='black')
    Frame_5_Widget_7.place(x=1, y=120)
    Frame_5_Widget_8 = CTkLabel(frame_5, textvariable=Frame_5_Widget_8_textvar , width=98 , font= fonte_b, fg_color='black') 
    Frame_5_Widget_8.place(x=101, y=120)

    Frame_5_Widget_9 = CTkLabel(frame_5, width=98 ,text='TOTAL', font= fonte_b, fg_color='white', text_color='black')
    Frame_5_Widget_9.place(x=1, y=150)
    Frame_5_Widget_10 = CTkLabel(frame_5, textvariable=Frame_5_Widget_10_textvar , width=98 , font= fonte_b, fg_color='white', text_color='black') 
    Frame_5_Widget_10.place(x=101, y=150)

#   FRAME 6 \\  Expositor faturamentos Mensal

if True:
    #Expositor Faturamento mensal

    #definições pré-widget
    faturamento_por_barbeiro_mensal = Obter_faturamento_por_barbeiro_mensal(periodo)
    total_mensal = Soma(faturamento_por_barbeiro_mensal)
    Frame_6_Widget_2_textvar = StringVar(value=f'R${faturamento_por_barbeiro_mensal[0]}')
    Frame_6_Widget_4_textvar = StringVar(value=f'R${faturamento_por_barbeiro_mensal[1]}')
    Frame_6_Widget_6_textvar = StringVar(value=f'R${faturamento_por_barbeiro_mensal[2]}')
    Frame_6_Widget_8_textvar = StringVar(value=f'R${faturamento_por_barbeiro_mensal[3]}')
    Frame_6_Widget_10_textvar = StringVar(value=f'R${total_mensal:.2f}')

    frame_6 = CTkFrame(tab_principal, width=201, height=180,border_width=1, border_color='white')
    frame_6.place(x=10, y=455)

    Frame_6_Widget_0 = CTkLabel(frame_6, width=197 ,text='Faturamento Mensal', font=fonte_b, fg_color='blue')
    Frame_6_Widget_0.place(x=2, y=2)

    Frame_6_Widget_1 = CTkLabel(frame_6, width=98 ,text=profissionais[0], font=fonte_b, fg_color='black')
    Frame_6_Widget_1.place(x=1, y=30)
    Frame_6_Widget_2 = CTkLabel(frame_6, textvariable=Frame_6_Widget_2_textvar ,width=98 ,font=fonte_b, fg_color='black')
    Frame_6_Widget_2.place(x=101, y=30)

    Frame_6_Widget_3 = CTkLabel(frame_6, width=98 ,text=profissionais[1], font= fonte_b, fg_color='black')
    Frame_6_Widget_3.place(x=1, y=60)
    Frame_6_Widget_4 = CTkLabel(frame_6, textvariable=Frame_6_Widget_4_textvar ,width=98, font= fonte_b, fg_color='black') 
    Frame_6_Widget_4.place(x=101, y=60)

    Frame_6_Widget_5 = CTkLabel(frame_6, width=98 ,text=profissionais[2], font= fonte_b, fg_color='black')
    Frame_6_Widget_5.place(x=1, y=90)
    Frame_6_Widget_6 = CTkLabel(frame_6, textvariable=Frame_6_Widget_6_textvar ,width=98 , font= fonte_b, fg_color='black') 
    Frame_6_Widget_6.place(x=101, y=90)
    
    Frame_6_Widget_7 = CTkLabel(frame_6, width=98 ,text=profissionais[3], font= fonte_b, fg_color='black')
    Frame_6_Widget_7.place(x=1, y=120)
    Frame_6_Widget_8 = CTkLabel(frame_6, textvariable=Frame_6_Widget_8_textvar , width=98 , font= fonte_b, fg_color='black') 
    Frame_6_Widget_8.place(x=101, y=120)

    Frame_6_Widget_9 = CTkLabel(frame_6, width=98 ,text='TOTAL', font= fonte_b, fg_color='white', text_color='black')
    Frame_6_Widget_9.place(x=1, y=150)
    Frame_6_Widget_10 = CTkLabel(frame_6, textvariable=Frame_6_Widget_10_textvar , width=98 , font= fonte_b, fg_color='white', text_color='black') 
    Frame_6_Widget_10.place(x=101, y=150)

#    FRAME 2 \\ Seção TreeView



if True:
    #Frame_2_Widget_0 \\ treeview
    frame_2 = CTkFrame(tab_principal)
    frame_2.place(relx=0.16, rely=0.1)

    tv = Treeview(frame_2, columns=colunas_database , show='headings', height=20, selectmode='browse')
    tv.grid(row=0, column=0)
    tv.column('id', minwidth=30, width=30, anchor=CENTER)
    tv.column('data', minwidth=50, width=50, anchor=CENTER)
    tv.column('profissional', minwidth=0, width=150, anchor=CENTER)
    tv.column('histórico', minwidth=310, width=310)
    tv.column('form pgmt', minwidth=0, width=80)
    tv.column('entrada', minwidth=0, width=70)
    tv.column('saída', minwidth=0, width=70)
    tv.column('hora', minwidth=50, width=50, anchor=CENTER)
    tv.heading('id', text='ID', anchor=CENTER)
    tv.heading('data', text='DATA')
    tv.heading('profissional', text='PROFISSIONAL')
    tv.heading('histórico', text='HISTÓRICO')
    tv.heading('form pgmt', text='FORM. PGMT')
    tv.heading('entrada', text='VALOR')
    tv.heading('saída', text='SAÍDA')
    tv.heading('hora', text='HORA')
    
    vs=CTkScrollbar(frame_2, command=tv.yview)
    tv.configure(yscrollcommand=vs.set)
    vs.grid(row=0, column=1, sticky='ns')

    LoadData()

    # seção de controle da treeview

    frame_2_widget_1 = CTkButton(frame_2, width=10, height=10 ,image=imagem_excluir, text='', hover=True, fg_color='white', command=lambda:DeleteTreeviewItem())
    frame_2_widget_1.grid(row=2, column=0, pady=(10, 0))

#   FRAME 3  \\ Seção Registros Entrada

if True:
    frame_3 = CTkFrame(tab_principal, border_width=1 ,border_color='white')
    frame_3.place(relx=0.77, rely=0.08)

    label_combobox_barbeiro = CTkLabel(frame_3, text='Profissional', font= fonte_b)
    label_combobox_barbeiro.grid(row=0, column=1, pady=(10, 0))
    combobox_barbeiro = CTkComboBox(frame_3,font= fonte, values=profissionais, fg_color='black')  
    combobox_barbeiro.grid(row=1, column=1, pady=(0, 15))
    combobox_barbeiro.set('PROFISSO.')

    label_10 = CTkLabel(frame_3, text='Serviços', font= fonte_b)
    label_10.grid(row=2, column=1)

    check1_control = StringVar()
    check1 = CTkCheckBox(frame_3, variable= check1_control,width=5,text='Corte',onvalue='CORTE', offvalue='', font= fonte)
    check1.grid(row=3, column=0, padx=(5, 0))
    check2_control = StringVar()
    check2 = CTkCheckBox(frame_3, variable= check2_control,width=5 ,text='Barba ', onvalue='BARBA', offvalue='', font= fonte)
    check2.grid(row=3, column=1, padx=(5, 0))
    check3_control = StringVar()
    check3 = CTkCheckBox(frame_3, variable= check3_control,width=5 ,text='Sobran.', onvalue='SOBRANCELHA', offvalue='', font= fonte)
    check3.grid(row=3, column=2, padx=(5))
    check4_control = StringVar()
    check4 = CTkCheckBox(frame_3, variable= check4_control,width=5 ,text=' Pig.', onvalue='PIGMENTAÇÃO', offvalue='', font= fonte)
    check4.grid(row=4, column=0, padx=(5), pady=(5,0))
    check5_control = StringVar()
    check5 = CTkCheckBox(frame_3, variable= check5_control,width=5 ,text='Luzes', onvalue='LUZES', offvalue='', font= fonte)
    check5.grid(row=4, column=1, padx=(5), pady=(5,0))
    check6_control = StringVar()
    check6 = CTkCheckBox(frame_3, variable= check6_control,width=5 ,text='Platin.  ', onvalue='PLATINADO', offvalue='', font= fonte)
    check6.grid(row=4, column=2, padx=(5), pady=(5,0))

    label_forma_pgmt = CTkLabel(frame_3, text='Forma de Pagamento', font= fonte_b)
    label_forma_pgmt.grid(row=5, column=1, pady=(15, 0))
    forma_pgmt = CTkComboBox(frame_3, font= fonte, values=form_pgmt_entrada, fg_color='black')  
    forma_pgmt.grid(row=6, column=1)
    forma_pgmt.set('FORM. PGMT.')

    label_entry_valor = CTkLabel(frame_3, text='Valor(R$):', font= fonte_b)
    label_entry_valor.grid(row=7, column=1, pady=(15, 0))
    entry_valor = CTkEntry(frame_3, font=fonte, fg_color='black')
    entry_valor.grid(row=8, column=1)

    frame_3_widget_14_img = CTkImage(light_image=Image.open(r'C:\Users\Customer\Desktop\nw_codigos\app\images\botao_bebida.png'), size=(30, 30 ))
    frame_3_widget_14 = CTkButton(frame_3, text='', width=10, height=10, image=frame_3_widget_14_img, fg_color='white', command=lambda:Botao_bebida_window())
    frame_3_widget_14.grid(row=8, column=2)

    botao_registrar = CTkButton(frame_3, state='normal',text='Registrar Entrada', command=lambda:RegistrarCorte(periodo), font= fonte_b)
    botao_registrar.grid(row=9, column=1, pady=(15, 10))  
    
#   Frame 4 \\ Seção Registro de Saída

if True:
    frame_4 = CTkFrame(tab_principal, border_width=1, border_color='white')
    frame_4.place(relx=0.77, rely=0.62)

    frame_4_widget_0 = CTkLabel(frame_4, text='Despesa', font= fonte_b)
    frame_4_widget_0.grid(row=0, column=0, padx=(85) ,pady=(5, 0))
    
    frame_4_widget_1 = CTkComboBox(frame_4, values=despesas, font=fonte, fg_color='black')
    frame_4_widget_1.grid(row=1, column=0, padx=(85) ,pady=(5, 0))
    frame_4_widget_1.set('DESPESA')

    frame_4_widget_2 = CTkLabel(frame_4, text='Forma de pagamento', font= fonte_b)
    frame_4_widget_2.grid(row=2, column=0 ,pady=(15, 0))

    frame_4_widget_3 = CTkComboBox(frame_4, font= fonte, values=form_pgmt_saida, fg_color='black')  
    frame_4_widget_3.grid(row=3, column=0, pady=(5, 0))
    frame_4_widget_3.set('FORM. PGMT.')

    frame_4_widget_4 = CTkLabel(frame_4, text='Valor(R$)', font= fonte_b)
    frame_4_widget_4.grid(row=4, column=0, padx=(85) ,pady=(15, 0))

    frame_4_widget_5 = CTkEntry(frame_4, font=fonte, fg_color='black')
    frame_4_widget_5.grid(row=5, column=0, padx=(85) ,pady=(5, 0))

    frame_4_widget_6 = CTkButton(frame_4, text='Registrar Saída', command=lambda:RegistrarSaida(), font= fonte_b)
    frame_4_widget_6.grid(row=6, column=0, pady=(10))  
    
# FRAME 7 \\ EXPOSITOR DO FATURAMENTOS DO DIA POR TIPO
    
    #pré- widgets
    
    #saldo = Obter_caixa(data, periodo)
    faturamento_dia_by_formpgmt = Get_faturamento_dia_by_formpgmt(ano, periodo, data) # [dinheiro, cartão, pix]
    caixa = Obter_caixa(data, periodo)

    #text var's

    frame_7_Widget_1_textvar = StringVar(value=f'R${faturamento_dia_by_formpgmt[0]}')
    frame_7_Widget_3_textvar = StringVar(value=f'R${faturamento_dia_by_formpgmt[1]}')
    frame_7_Widget_5_textvar = StringVar(value=f'R${faturamento_dia_by_formpgmt[2]}')
    frame_7_Widget_12_textvar = StringVar(value=f'R${faturamento_dia_by_formpgmt[3]}')   
    frame_7_Widget_8_textvar = StringVar(value=f'R${caixa}')

    #frame

    frame_7 = CTkFrame(tab_principal, width=790, height=100, border_width=1, border_color='white')
    frame_7.place(relx=0.175, rely=0.8)
    
    #widget's

    frame_7_Widget_0 = CTkLabel(frame_7, width=90 , text=' DINHEIRO ', font=fonte_b, fg_color='blue')
    frame_7_Widget_0.place(x=20, y=10)
    frame_7_Widget_1 = CTkLabel(frame_7, width=90 ,textvariable=frame_7_Widget_1_textvar, font=fonte_b, fg_color='black')
    frame_7_Widget_1.place(x=110, y=10)

    frame_7_Widget_2 = CTkLabel(frame_7, width=90 , text=' DÉBITO ', font=fonte_b, fg_color='blue')
    frame_7_Widget_2.place(x=210, y=10)
    frame_7_Widget_3 = CTkLabel(frame_7, width=90 ,textvariable=frame_7_Widget_3_textvar, font=fonte_b, fg_color='black')
    frame_7_Widget_3.place(x=300, y=10)

    frame_7_Widget_4 = CTkLabel(frame_7, width=90 , text=' CRÉDITO ', font=fonte_b, fg_color='blue')
    frame_7_Widget_4.place(x=400, y=10)
    frame_7_Widget_5 = CTkLabel(frame_7, width=90 ,textvariable=frame_7_Widget_5_textvar, font=fonte_b, fg_color='black')
    frame_7_Widget_5.place(x=490, y=10)

    frame_7_Widget_11 = CTkLabel(frame_7, width=90 , text=' PIX ', font=fonte_b, fg_color='blue')
    frame_7_Widget_11.place(x=590, y=10)
    frame_7_Widget_12 = CTkLabel(frame_7, width=90 ,textvariable=frame_7_Widget_12_textvar, font=fonte_b, fg_color='black')
    frame_7_Widget_12.place(x=680, y=10)

    frame_7_Widget_6 = CTkButton(frame_7, text='Fechar Caixa', command=lambda:FecharCaixa(frame_7_Widget_10.get()) ,font= fonte_b)
    frame_7_Widget_6.place(x=570, y=60)  

    frame_7_Widget_7 = CTkLabel(frame_7, width=90 , text='    CAIXA     ', font=fonte_c, fg_color='blue')
    frame_7_Widget_7.place(x=20, y=60)
    frame_7_Widget_8 = CTkLabel(frame_7, width=130 ,textvariable=frame_7_Widget_8_textvar, font=fonte_c, fg_color='black', text_color='green')
    frame_7_Widget_8.place(x=155, y=60)

    frame_7_Widget_9 = CTkLabel(frame_7, text='R$' ,font=fonte_c)
    frame_7_Widget_9.place(x=325, y=60)
    frame_7_Widget_10 = CTkEntry(frame_7, width=118 ,font=fonte, fg_color='black')
    frame_7_Widget_10.place(x=363, y=60)


# FRAME 8 \\ EXPOSITOR DE PERIODO

    #textvars's 

    



janela_main.mainloop()

