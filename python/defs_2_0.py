import pandas as pd
from tkinter import *
from datetime import datetime, date
import openpyxl as op
from CTkMessagebox import CTkMessagebox
from defsofdefs import *
from time import sleep

profissionais = ['VITOR', 'FERNANDO', 'FREE1', 'FREE2']
arquivo_name = 'defs.py'
despesas = ['CONSUMÍVEIS', 'ALUGUEL', 'ÁGUA', 'LUZ', 'INTERNET','FUNDOS', 'ESTORNO', 'OUTRO']
form_pgmt_saida = ['DINHEIRO', 'BANCO', 'CARTÃO']
form_pgmt_entrada = ['DINHEIRO', 'PIX QR', 'DÉBITO', 'CRÉDITO','PIX CHAVE','MENSAL']
valor_limite_saida = 9999
valor_limite_entrada = 9999


def Lin():
    print('-------------------------------')

def Check_datas(data):
    lista_datas = Get_lista_datas_semana()
    if data in lista_datas:
        return True
    else:
        return False

def Get_lista_datas_semana():
    with open('txts\datas_semana.txt', 'r') as a:
        return a.readline().split(';')

def Update_lista_datas_semana(data, ano):
    def Get_first_week_date(data, ano, weekday_indice):#identificar da data da ultima segunda-feira seja ela no mes ou ano passado
        try:
            data_splited = data.split('-')
            day = int(data_splited[0])
            month = int(data_splited[1])
            year = int(ano)
            if weekday_indice == 0:
                return data
            else:#para o caso de o dia em questão não for segunda
                day = int(day)
                month = int(month)
                for i in range(weekday_indice+1):
                    if day < 1:
                        if month == 1:
                            day = 31
                            month = 12
                            ano = str(int(ano)-1)
                        else:    
                            day = DiasMes(month-1, ano)
                            month -= 1
                    if i == weekday_indice:
                        return [int(ano), month, day]
                    day -= 1
        except:
            print('Erro na subfunção Get_first_week_date da função Update_lista_semanal, arq defs.py.')
    try:
        day, month = data.split('-')
        data_datetime = date(int(ano),int(month),int(day))
        weekday_indice = data_datetime.weekday()
        if weekday_indice != 0:
            year, month, day = Get_first_week_date(data, ano, weekday_indice)
        else:
            day = data_datetime.day
            month = data_datetime.month
            year = data_datetime.year
        #formar a lista de datas desde a ultima segunda feira
        lista_datas = []
        dias_mes = DiasMes(month, ano[2:])
        for i in range(7):#loop que acrescenta as datas da semana na lista
            if day > dias_mes:#condição que recunha a virada do mes
                day = 1
                if month < 12:
                    month += 1
                else:
                    month=1
                    year += 1
            variable_data = Zero_adder(str(day))+'-'+Zero_adder(str(month))
            lista_datas.append(variable_data)
            day += 1
        with open('txts\datas_semana.txt', 'w') as a:
            joined_list=';'.join(lista_datas)
            a.write(joined_list)
            print(f'Lista de datas da semana foram atualizadas: {lista_datas}')
    except:
        print('Erro na função update_datas_semana no arq def.py.')

def Obter_faturamento_diario_by_barbeiro(wb, data, periodo):
    try:
        #abrindo a sheet do dt
        sheet = wb[periodo]
        rows = list(sheet.values)
        #identificando as rows certas
        profissional1 = profissional2 = profissional3 = profissional4 = 0
        for row in rows:
            #filtrando data
            if row[1] == data:
                #tratando a lista
                row = list(row)
                if '+' in row[7]:
                    row[7]=Soma(row[7].split(' + '))
                #filtrando o profissional
                if row[2] == profissionais[0]:
                    profissional1 += float(row[7])
                elif row[2] == profissionais[1]:
                    profissional2 += float(row[7])
                elif row[2] == profissionais[2]:
                    profissional3 += float(row[7])
                elif row[2] == profissionais[3]:
                    profissional4 += float(row[7])
        return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']
    except:
        print('Erro ao calcular o faturamento diário por barbeiro!')
        return ['0,00', '0,00', '0,00', '0,00']

def Obter_faturamento_semanal_by_barbeiro(data, periodo):
    def Get_meses_lista_datas_semana(lista_datas):
        #formar lista dos meses diferentes separando-os dos dias e tratando-os
        meses = []
        last_mes = ''
        for data in lista_datas:
            data_splited = data.split('-')
            mes = data_splited[1]
            if last_mes != mes:
                meses.append(mes)
            last_mes = mes
        return meses # [atual] | [antigo, atual]
    #try:
    #definições importantes
    data_splited = data.split('-')
    mes_data = data_splited[1]
    periodo_splited = periodo.split('-')
    ano = '20' + periodo_splited[1]
    ano_abrev = periodo_splited[1]
    #obtenção das datas da semana
    if Check_datas(data) == False:#check se as datas da semana são válidas para a data atual
        Update_lista_datas_semana(data, ano)#atualiza as datas da semana
        lista_datas = Get_lista_datas_semana()
    else:
        lista_datas = Get_lista_datas_semana() 
    #calculando o faturamento de cada profissional no(s) periodo(s)
    lista_meses = Get_meses_lista_datas_semana(lista_datas)#o len dessa lista def se as datas da semana estão contidos somente no periodo mensal atual ou se tem mais de um
    profissional1 = profissional2 = profissional3 = profissional4 = 0
    for i, mes in enumerate(lista_meses[::-1], start=1):#passando pelo(s) database(s) do mais novo pro mais velho
        if mes <= mes_data:
            if i == 2:#analisando apenas o segundo mes do for
                if mes == '12':
                    ano = str(int(ano)-1)
                    print(f'troquei o ano: {ano}')
            #abrir o data base do periodo mensal da instancia
            wb = op.load_workbook(fr'excell\nw_barbearia_{ano}.xlsx')
            ws = wb[f'{mes}-{ano_abrev}']
            rows = list(ws.values)
            for row in rows:
                #filtrando data
                if row[1] in lista_datas:
                    row = list(row)
                    if '+' in row[7]:
                        row[7]=Soma(row[7].split(' + '))
                    #filtrando o profissional
                    if row[2] == profissionais[0]:
                        profissional1 += float(row[7])
                    elif row[2] == profissionais[1]:
                        profissional2 += float(row[7])
                    elif row[2] == profissionais[2]:
                        profissional3 += float(row[7])
                    elif row[2] == profissionais[3]:
                        profissional4 += float(row[7])
    return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']
    #except:
        #print(f'Erro durante a execução da função Obter_faturamento_por_barbeiro_semanal no arquivo {arquivo_name}. lista retornada: [''0.00'', ''0.00'', ''0.00'', ''0.00'']')
        #return ['0.00', '0.00', '0.00', '0.00']

def Obter_faturamento_mensal_by_barbeiro(wb, periodo):
    try:
        #abrindo a database
        sheet = wb[periodo]
        rows = list(sheet.values)
        #identificando as rows certas
        profissional1 = profissional2 = profissional3 = profissional4 = 0
        for row in rows:
            row = list(row)
            if '+' in row[7]:
                row[7]=Soma(row[7].split(' + '))
            #filtrando o profissional
            if row[2] == profissionais[0]:
                profissional1 += float(row[7])
            elif row[2] == profissionais[1]:
                profissional2 += float(row[7])
            elif row[2] == profissionais[2]:
                profissional3 += float(row[7])
            elif row[2] == profissionais[3]:
                profissional4 += float(row[7])
        return [f'{profissional1:.2f}', f'{profissional2:.2f}', f'{profissional3:.2f}', f'{profissional4:.2f}']
    except:
        print('Erro ao calcular o faturamento mensal por barbeiro! lista retornada: [''0.00'', ''0.00'', ''0.00'', ''0.00'']')
        return ['0,00', '0,00', '0,00', '0,00']

def Obter_total_entrada_dinheiro(data, periodo):
    try:
        ano = '20'+periodo[3:]
        #abrindo o databse
        book = op.load_workbook(fr'excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        rows = list(sheet.values)
        total = 0
        for row in rows:
            if row[1] == data:
                if row[6] == 'DINHEIRO' and row[8]==None:
                    total += float(row[7])
        return total
    except:
        print('Erro durante a execução da função Obter_total_dinheiro (defs.py). Valor retornado: 00.00')
        return float('00.00')


def Obter_total_saida_dinheiro(data, periodo):
    try:
        ano = '20'+periodo[3:] 
        #abrindo o databse
        book = op.load_workbook(fr'excell\nw_barbearia_{ano}.xlsx')
        sheet = book[periodo]
        rows = list(sheet.values)
        total = 0
        for row in rows:
            if row[1] == data:
                if row[6] == 'DINHEIRO' and row[7]==None:
                    total += float(row[8])
        return total
    except:
        print('Erro durante a execução da função Obter_total_saida_dinheiro (defs.py). Valor retornado: 00.00')
        return float('00.00')

def Obter_ultimo_caixa():
    try:
        arquivo = open(r'txts\caixa.txt', 'r')
        saldo = float(arquivo.readline())
        return saldo
    except:
        print('Erro durante a execução da função Obter_ultimo_caixa (defs.py). Valor retornado: 00.00')
        return float(f'00.00')

def Obter_caixa(data, periodo):
    try:
        ultimo_caixa = Obter_ultimo_caixa()
        total_entrada_dinheiro = Obter_total_entrada_dinheiro(data, periodo)
        total_saida_dinheiro = Obter_total_saida_dinheiro(data, periodo)
        caixa = ultimo_caixa + total_entrada_dinheiro - total_saida_dinheiro
        return caixa
    except:
        print('Ocorreu um erro durante a execução da função Obter_caixa (defs.py). Valor retornado: 00.00')
        return float(f'00.00')

def GetPeriodo(colunas_database):
    #PE0: pode ocorrer que haja a criação de wb's ou wb's com nomes repetidos
    try:
        def CheckPeriodo_anual():
            try:
                #definindo periodos anuais setados e do computador
                arquivo = open(r'txts\periodo_anual.txt', 'r')
                periodo_anual_setado = arquivo.readline()
                periodo_anual_computador = date.today().year
                #comparando os periodos anuais do txt e do computador
                if periodo_anual_computador > int(periodo_anual_setado):
                    print(f'Novo periodo anual identificado:{periodo_anual_computador}')
                    return [False, periodo_anual_computador]
                else:
                    return [True, periodo_anual_setado]
            except:
                print('Ocorreu um ERRO na subfunção CheckPeriodo_mensal da função GetPeriodo em defs.py. Periodo anual retornado: 0000')
                return [True, '0000']

        def ChangePeriodo_anual(colunas_database, periodo_anual_computador):
            try:    
                #criar um novo Wb
                wb = op.Workbook()
                mes_computador = Zero_adder(date.today().month)
                novo_periodo = f'{mes_computador}-{str(periodo_anual_computador)[2:]}'
                sheet = wb.active
                sheet.title = novo_periodo
                sheet.append(colunas_database)
                wb.save(rf'excell\nw_barbearia_{periodo_anual_computador}.xlsx')
                print(f'Novo arquivo excell criado referente ao novo periodo anual {periodo_anual_computador} juntamento com o periodo mensal {mes_computador}')
                #reescrever o conteudo do arquivo periodo_anual.txt
                arquivo = open(r'txts\periodo_anual.txt', 'w')
                arquivo.write(str(periodo_anual_computador))
                print(f'Novo periodo anual setado: {periodo_anual_computador}')
                #finalizando a função
                return novo_periodo
            except:
                print('Ocorreu um ERRO na subfunção ChangePeriodo_anual da função GetPeriodo em defs.py. Periodo retornado: 00-00')
                return '00-00'
        
        def CheckPeriodo_mensal():
            try:
                #definindo periodos mensais setados e do computador
                arquivo = open(r'txts\periodo_mensal.txt', 'r')
                periodo_mensal_setado = arquivo.readline()
                periodo_mensal_computador = Zero_adder(date.today().month)
                #comparando os periodos mensais do txt e do computador
                if int(periodo_mensal_computador) - int(periodo_mensal_setado) != 0:
                    return [False, periodo_mensal_computador]
                else:
                    return [True, periodo_mensal_setado]
            except:
                print('Ocorreu um ERRO na subfunção CheckPeriodo_mensal da função GetPeriodo em defs.py. Periodo mensal retornado: 00')
                return [True, '00']


        def ChangePeriodo_mensal(colunas_database, periodo_mensal_computador, periodo_anual_setado):
            try:
                novoperiodo = f'{(periodo_mensal_computador)}-{periodo_anual_setado[2:]}'
                #criando uma nova sheet referente ao novo periodo mensal
                wb = op.load_workbook(rf'excell\nw_barbearia_{periodo_anual_setado}.xlsx')#abrindo o db
                wb.create_sheet(novoperiodo)#criando nova sheet correspondente ao novo periodo mensalexcell\nw_barbearia_2024.xlsx
                #acrescentando os titulo das colunas do database
                sheet = wb[novoperiodo]
                sheet.append(colunas_database)
                wb.save(rf'excell\nw_barbearia_{periodo_anual_setado}.xlsx')
                print(f'Uma nova sheet com o nome {novoperiodo} foi criada no arquivo nw_barbearia_{periodo_anual_setado}.xlsx')
                #atualizar no txt
                arquivo = open(r'txts\periodo_mensal.txt', 'w')
                arquivo.write(str(periodo_mensal_computador))
                print(f'Novo periodo mensal setado: {periodo_mensal_computador}')
                return novoperiodo
            except:
                print('Ocorreu um erro na subfunção ChangePeriodo_mensal da função GetPeriodo em defs.py. Periodo retornado: 00-00')
                return '00-00'
        
        check_periodo_anual = CheckPeriodo_anual()
        if check_periodo_anual[0] == False:
            periodo_novo = ChangePeriodo_anual(colunas_database, check_periodo_anual[1])
            return periodo_novo
        else:#se o periodo anual nao mudou, o pg ira agr checkar o periodo mensal
            check_periodo_mensal = CheckPeriodo_mensal()
            if check_periodo_mensal[0]==False:
                periodo_novo = ChangePeriodo_mensal(colunas_database, check_periodo_mensal[1], check_periodo_anual[1])
                return periodo_novo
            else:#para o caso de o periodo mensal nao ter mudado
                periodo = f'{check_periodo_mensal[1]}-{check_periodo_anual[1][2:]}'
                return periodo
    except:
        print('Ocorreu um erro na função GetPeriodo em defs.py. Periodo retornado: 00-00')
        return '00-00'

def GetData():
    dia = Zero_adder(str(date.today().day))
    mes = Zero_adder(str(date.today().month))
    data = f'{dia}-{mes}'
    return data

def GetHora():
    hr = str(datetime.now())[11:16]
    return hr

def GetLastId(wb, periodo, lista):
    sheet = wb[periodo]
    rows = list(sheet.values)
    if rows[-1] == lista:#p\ o caso de nao haver movimentações
        return 0
    return rows[-1][0]

def Check_entrada(profissional, meiopgmt, entrada, servicos, bebidas, produtos):#checkar inputs da area de registro de entradas
    algumacoisa = False
    if profissional != '':#verificar se o usuario nao vai colocar um nome inexistente ou errado
        if profissional not in profissionais:
            print('Opção escolhida ''profissional'' não foi aceita')
            msg = 'Verifique a opção Profissional e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False
    if '+' in meiopgmt:#caso haja 2 form de pgmt
        meiospgmt = meiopgmt.split(' + ')
        meiopgmt1 = meiospgmt[0]
        meiopgmt2 = meiospgmt[1]
        entradas = entrada.split(' + ')
        entrada1 = entradas[0]
        entrada2 = entradas[1]
        if meiopgmt1 not in form_pgmt_entrada:
            print('Opção escolhida ''forma de pagamento 1'' não foi aceita')
            msg = 'Verifique a opção Forma de Pagamento 1 e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False
        if entrada1 == '' or str(entrada1).replace('.','').isnumeric()==False or float(entrada1) > valor_limite_entrada or float(entrada1) <= 0:
            print('Opção escolhida ''Valor 1'' não foi aceita')
            msg = 'Verifique a opção Valor 1 e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False 
        if meiopgmt2 not in form_pgmt_entrada:
            print('Opção escolhida ''forma de pagamento 2'' não foi aceita')
            msg = 'Verifique a opção Forma de Pagamento 2 e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False
        if entrada2 == '' or str(entrada2).replace('.','').isnumeric()==False or float(entrada2) > valor_limite_entrada or float(entrada2) <= 0:
            print('Opção escolhida ''Valor 2'' não foi aceita')
            msg = 'Verifique a opção Valor 2 e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False
    else:#caso haja so 1 form pgmt
        if meiopgmt not in form_pgmt_entrada:
            print('Opção escolhida ''forma de pagamento'' não foi aceita')
            msg = 'Verifique a opção Forma de Pagamento e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False
        if entrada == '' or str(entrada).replace('.','').isnumeric()==False or float(entrada) > valor_limite_entrada or float(entrada) <= 0:
            print('Opção escolhida ''Valor'' não foi aceita')
            msg = 'Verifique a opção Valor e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False

    if servicos != '':
        algumacoisa = True
    if bebidas != '':
        algumacoisa = True
    if produtos != '':
        algumacoisa = True

    if algumacoisa == True:
        return True
    else:
        print('Nenhum serviço, bebida ou produto selecionado.')
        msg = 'Selecione um serviço, bebida ou produto e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
        return False

def Check_saida(despesa, meiopgmt, saida):#checkar inuts da area de registro de saidas
    if despesa not in despesas:
        print('Opção escolhida ''despesa'' não foi aceita')
        msg = 'Verifique a opção Despesa e tente novamente!'.upper()
        CTkMessagebox(title='Não foi possível Registrar a Saída', message=msg, icon='cancel')
        return False
    if meiopgmt not in form_pgmt_saida:
            print('Opção escolhida ''forma de pagamento'' não foi aceita')
            msg = 'Verifique a opção Forma de Pagamento e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False
    if saida == '' or str(saida).replace('.','').isnumeric()==False or float(saida) > valor_limite_saida or float(saida) <= 0:
            print('Opção escolhida ''Valor'' não foi aceita')
            msg = 'Verifique a opção Valor e tente novamente!'.upper()
            CTkMessagebox(title='Não foi possível Registrar a Entrada', message=msg, icon='cancel')
            return False
    return True
        
def ObterListaProfissionais():
    try:
        with open(r'txts\profissionais.txt', 'r') as a:
            linhas = a.readline().split(';')
            profissionais = []
            for linha in linhas:
                profissionais.append(linha)
            return profissionais

        #returnprofissionais
    except:
        print('Erro ao capturar a lista de profissionais!')
        return ['###', '###', '###', '###']

def Soma(lista):
    try:
        soma = 0
        for item in lista:
            soma+=float(item)
        return soma
    except:
        print('Houve um erro na função Soma')

def CloseApp(master):
    print('fechando app...')
    master.destroy()
    print('App fechado!')

def DiasMes(mes, ano):
    try:   
        month = int(mes)
        if month == 2:#se é o mês fevereiro
            ano = int('20'+ano)
            if ano % 4 == 0:#se for bissexto
                return 29
            else:
                return 28
        if month % 2 != 0:#se o mes for ímpar
            if month <= 7:
                return 31
            else:
                return 30
        if month % 2 == 0:#se o mes for par
            if month >= 8:
                return 31
            else:
                return 30
    except:
        print(f'Erro durante a execução da função DiasMes no arquivo {arquivo_name}.Valor retornado: 0')
        return 0

def Get_faturamento_dia_by_formpgmt(wb, periodo, data):
    try:
        #abrindo a database
        sheet = wb[periodo]
        rows = list(sheet.values)
        dinheiro = debito = credito = pix = 0
        if len(rows) == 1:
            return ['0.00', '0.00','0.00', '0.00']
        #identificando as rows certas
        for row in rows[1:]:
            #filtrando data
            if row[1] ==data:
                if '+' in row[6]:#rows com form de pgmt composta
                    #decompondo a row em 2
                    lista_forms_pgmt = row[6].split(' + ')
                    lista_valores = row[7].split(' + ')
                    row1 = f'{lista_forms_pgmt[0]} + {lista_valores[0]}'
                    row2 = f'{lista_forms_pgmt[1]} + {lista_valores[1]}'
                    rowfinal = [row1, row2]
                    for item in rowfinal:
                        item = item.split(' + ')
                        if 'DINHEIRO' == item[0]:
                            dinheiro += float(item[1])
                        elif 'DÉBITO' == item[0]:
                            debito += float(item[1])
                        elif 'CRÉDITO' == item[0]:
                            credito += float(item[1])        
                        elif 'PIX QR' == item[0]:
                            pix += float(item[1])
                else:#rows com form de pgmt unica
                    if 'DINHEIRO' == row[6]:
                        if row[7] != None:
                            dinheiro += float(row[7])
                    elif 'DÉBITO' == row[6]:
                        if row[7] != None:
                            debito += float(row[7])
                    elif 'CRÉDITO' == row[6]:
                        if row[7] != None:
                            credito += float(row[7])        
                    elif 'PIX QR' == row[6]:
                        if row[7] != None:
                            pix += float(row[7])
        return [dinheiro, debito, credito, pix]
    except:
        print('Erro na função Get_faturamento_dia_by_formpgmt. lista retornada: [''0.00'', ''0.00'',''0.00'', ''0.00'']')
        return ['0.00', '0.00','0.00', '0.00']

def Show_success_msgbox(master, msg):
    box = CTkMessagebox(master, title='', message=msg, icon='check', fade_in_duration=1, button_hover_color='white')
    box.destroy()

def List_change_nothing_to_something(lista, character):
    try:
        new_lista = []
        for item in lista:
            if item == None or item == '':
                new_lista.append(character)
            else:
                new_lista.append(item)
        return new_lista
    except:
        print('Erro na função List_change_nothing_to_something, arq def.py. retornando a mesma lista recebida cm parametro')
        return lista